from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from app.models import GenerateRequest, ChatRequest, BOPData, UnifiedChatRequest, UnifiedChatResponse
from app.llm_service import generate_bop_from_text, modify_bop, unified_chat
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import json

app = FastAPI(title="Backend API", version="1.0.0")

# CORS 설정
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:5174", "http://localhost:5175"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
async def root():
    return {"message": "Backend Ready"}


@app.get("/health")
async def health():
    return {"status": "ok"}


@app.post("/api/generate")
async def generate_bop(req: GenerateRequest) -> BOPData:
    """
    사용자 입력을 받아 Gemini API를 통해 BOP를 생성합니다.
    """
    try:
        # LLM 서비스를 통해 BOP 생성
        bop_dict = await generate_bop_from_text(req.user_input)

        # Pydantic 모델로 validation
        bop_data = BOPData(**bop_dict)

        return bop_data

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"BOP 생성 실패: {str(e)}")


@app.post("/api/chat")
async def chat(req: ChatRequest) -> BOPData:
    """
    현재 BOP와 사용자 메시지를 받아 수정된 BOP를 반환합니다.
    """
    try:
        # current_bop을 dict로 변환
        current_bop_dict = req.current_bop.model_dump()

        # LLM 서비스를 통해 BOP 수정
        updated_bop_dict = await modify_bop(current_bop_dict, req.message)

        # Pydantic 모델로 validation
        updated_bop = BOPData(**updated_bop_dict)

        return updated_bop

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"BOP 수정 실패: {str(e)}")


@app.post("/api/chat/unified")
async def unified_chat_endpoint(req: UnifiedChatRequest) -> UnifiedChatResponse:
    """
    통합 채팅 엔드포인트: BOP 생성, 수정, QA를 모두 처리합니다.
    """
    try:
        # current_bop을 dict로 변환 (있는 경우)
        current_bop_dict = req.current_bop.model_dump() if req.current_bop else None

        # LLM 서비스를 통해 통합 처리
        response_data = await unified_chat(req.message, current_bop_dict)

        # bop_data가 있으면 Pydantic 모델로 validation
        bop_data = None
        if "bop_data" in response_data:
            bop_data = BOPData(**response_data["bop_data"])

        # UnifiedChatResponse 반환
        return UnifiedChatResponse(
            message=response_data["message"],
            bop_data=bop_data
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Chat 실패: {str(e)}")


@app.post("/api/export/excel")
async def export_excel(bop: BOPData):
    """
    BOP 데이터를 Excel 파일로 내보냅니다.

    Sheet 1: Overview (Project Info + Resource Masters)
    Sheet 2: BOP Processes
    """
    try:
        wb = Workbook()

        # ==================== Sheet 1: Overview ====================
        ws_overview = wb.active
        ws_overview.title = "Overview"

        # Project Info
        ws_overview.append(["Project Information"])
        ws_overview['A1'].font = Font(bold=True, size=14)
        ws_overview.append(["Project Title", bop.project_title])
        ws_overview.append(["Target UPH", bop.target_uph])
        ws_overview.append([])

        # Equipment Master
        ws_overview.append(["Equipment Master"])
        ws_overview[f'A{ws_overview.max_row}'].font = Font(bold=True, size=12)
        ws_overview.append(["Equipment ID", "Name", "Type"])
        header_row = ws_overview.max_row
        for col in range(1, 4):
            ws_overview.cell(header_row, col).font = Font(bold=True)
            ws_overview.cell(header_row, col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for eq in bop.equipments:
            ws_overview.append([
                eq.equipment_id,
                eq.name,
                eq.type
            ])

        ws_overview.append([])

        # Worker Master
        ws_overview.append(["Worker Master"])
        ws_overview[f'A{ws_overview.max_row}'].font = Font(bold=True, size=12)
        ws_overview.append(["Worker ID", "Name", "Skill Level"])
        header_row = ws_overview.max_row
        for col in range(1, 4):
            ws_overview.cell(header_row, col).font = Font(bold=True)
            ws_overview.cell(header_row, col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for worker in bop.workers:
            ws_overview.append([
                worker.worker_id,
                worker.name,
                worker.skill_level or "-"
            ])

        ws_overview.append([])

        # Material Master
        ws_overview.append(["Material Master"])
        ws_overview[f'A{ws_overview.max_row}'].font = Font(bold=True, size=12)
        ws_overview.append(["Material ID", "Name", "Unit"])
        header_row = ws_overview.max_row
        for col in range(1, 4):
            ws_overview.cell(header_row, col).font = Font(bold=True)
            ws_overview.cell(header_row, col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for material in bop.materials:
            ws_overview.append([
                material.material_id,
                material.name,
                material.unit
            ])

        # ==================== Sheet 2: BOP Processes ====================
        ws_bop = wb.create_sheet("BOP Processes")

        # Header
        ws_bop.append([
            "Process ID", "Name", "Description", "Cycle Time (s)", "Parallel Count",
            "Effective Time (s)", "Location (X,Y,Z)", "Predecessors", "Successors",
            "Equipments", "Workers", "Materials"
        ])
        header_row = 1
        for col in range(1, 13):
            ws_bop.cell(header_row, col).font = Font(bold=True)
            ws_bop.cell(header_row, col).fill = PatternFill(start_color="4a90e2", end_color="4a90e2", fill_type="solid")
            ws_bop.cell(header_row, col).font = Font(bold=True, color="FFFFFF")
            ws_bop.cell(header_row, col).alignment = Alignment(horizontal="center", vertical="center")

        # Data
        for process in bop.processes:
            # 병렬 라인 수만큼 row 생성
            for parallel_idx in range(process.parallel_count):
                # 리소스 요약 생성
                equipments_summary = []
                workers_summary = []
                materials_summary = []

                if hasattr(process, 'resources') and process.resources:
                    for resource in process.resources:
                        if resource.resource_type == "equipment":
                            eq = next((e for e in bop.equipments if e.equipment_id == resource.resource_id), None)
                            eq_name = eq.name if eq else resource.resource_id
                            equipments_summary.append(f"{eq_name} (x{resource.quantity})")
                        elif resource.resource_type == "worker":
                            worker = next((w for w in bop.workers if w.worker_id == resource.resource_id), None)
                            worker_name = worker.name if worker else resource.resource_id
                            workers_summary.append(f"{worker_name} (x{resource.quantity})")
                        elif resource.resource_type == "material":
                            mat = next((m for m in bop.materials if m.material_id == resource.resource_id), None)
                            mat_name = mat.name if mat else resource.resource_id
                            unit = mat.unit if mat else "ea"
                            materials_summary.append(f"{mat_name} ({resource.quantity}{unit})")

                equipments_str = ", ".join(equipments_summary) if equipments_summary else "-"
                workers_str = ", ".join(workers_summary) if workers_summary else "-"
                materials_str = ", ".join(materials_summary) if materials_summary else "-"

                # 첫 번째 라인만 전체 정보 표시
                if parallel_idx == 0:
                    # Safe handling of predecessor/successor IDs
                    predecessor_str = "-"
                    if hasattr(process, 'predecessor_ids') and process.predecessor_ids:
                        predecessor_str = ", ".join(process.predecessor_ids)

                    successor_str = "-"
                    if hasattr(process, 'successor_ids') and process.successor_ids:
                        successor_str = ", ".join(process.successor_ids)

                    ws_bop.append([
                        process.process_id,
                        process.name,
                        process.description,
                        process.cycle_time_sec,
                        process.parallel_count,
                        process.effective_cycle_time_sec,
                        f"({process.location.x}, {process.location.y}, {process.location.z})",
                        predecessor_str,
                        successor_str,
                        equipments_str,
                        workers_str,
                        materials_str
                    ])
                else:
                    # 병렬 라인은 간략하게 표시
                    ws_bop.append([
                        f"{process.process_id}-#{parallel_idx + 1}",
                        f"(병렬 라인 #{parallel_idx + 1})",
                        "-",
                        "-",
                        "-",
                        "-",
                        f"({process.location.x}, {process.location.y}, {process.location.z + parallel_idx * 5})",
                        "-",
                        "-",
                        "-",
                        "-",
                        "-"
                    ])

        # Adjust column widths
        ws_overview.column_dimensions['A'].width = 20
        ws_overview.column_dimensions['B'].width = 30
        ws_overview.column_dimensions['C'].width = 20

        ws_bop.column_dimensions['A'].width = 12
        ws_bop.column_dimensions['B'].width = 25
        ws_bop.column_dimensions['C'].width = 35
        ws_bop.column_dimensions['D'].width = 15
        ws_bop.column_dimensions['E'].width = 15
        ws_bop.column_dimensions['F'].width = 15
        ws_bop.column_dimensions['G'].width = 20
        ws_bop.column_dimensions['H'].width = 20
        ws_bop.column_dimensions['I'].width = 20
        ws_bop.column_dimensions['J'].width = 30
        ws_bop.column_dimensions['K'].width = 25
        ws_bop.column_dimensions['L'].width = 35

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Generate filename
        safe_filename = "".join(c for c in bop.project_title if c.isalnum() or c in (' ', '_', '-')).strip()
        if not safe_filename:
            safe_filename = "BOP"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={safe_filename}.xlsx"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel export 실패: {str(e)}")


def _get_color_for_equipment_type(equipment_type: str) -> str:
    """Equipment type에 따라 색상 반환"""
    color_map = {
        "robot": "#4a90e2",
        "machine": "#ff6b6b",
        "manual_station": "#50c878"
    }
    return color_map.get(equipment_type, "#888888")


@app.post("/api/export/3d")
async def export_3d(bop: BOPData):
    """
    BOP 데이터를 3D 시각화용 JSON으로 내보냅니다.
    Process 위치 + Resource 상대 위치로 실제 배치 계산
    """
    try:
        export_data = {
            "project_title": bop.project_title,
            "target_uph": bop.target_uph,
            "processes": [],
            "resources": []
        }

        # Process 정보
        for process in bop.processes:
            # 병렬 라인 수만큼 복제
            for parallel_idx in range(process.parallel_count):
                # 병렬 라인은 z축으로 5m씩 이동
                process_z_offset = parallel_idx * 5

                process_obj = {
                    "process_id": process.process_id if parallel_idx == 0 else f"{process.process_id}-#{parallel_idx + 1}",
                    "name": process.name,
                    "parallel_index": parallel_idx,
                    "position": {
                        "x": process.location.x,
                        "y": process.location.y,
                        "z": process.location.z + process_z_offset
                    },
                    "size": {
                        "width": 4.0,
                        "height": 2.0,
                        "depth": 3.0
                    },
                    "color": "#e0e0e0"
                }
                export_data["processes"].append(process_obj)

                # 이 공정의 리소스들 배치
                if hasattr(process, 'resources') and process.resources:
                    for resource in process.resources:
                        # 실제 위치 = Process 위치 + Resource 상대 위치
                        actual_x = process.location.x + resource.relative_location.x
                        actual_y = process.location.y + resource.relative_location.y
                        actual_z = process.location.z + process_z_offset + resource.relative_location.z

                        resource_obj = {
                            "resource_id": resource.resource_id,
                            "resource_type": resource.resource_type,
                            "process_id": process.process_id,
                            "parallel_index": parallel_idx,
                            "quantity": resource.quantity,
                            "role": resource.role or "",
                            "position": {
                                "x": actual_x,
                                "y": actual_y,
                                "z": actual_z
                            }
                        }

                        # 리소스 타입별 추가 정보
                        if resource.resource_type == "equipment":
                            equipment = next((eq for eq in bop.equipments if eq.equipment_id == resource.resource_id), None)
                            if equipment:
                                resource_obj["name"] = equipment.name
                                resource_obj["equipment_type"] = equipment.type
                                resource_obj["color"] = _get_color_for_equipment_type(equipment.type)
                                resource_obj["size"] = {
                                    "width": 1.0,
                                    "height": 2.0,
                                    "depth": 1.0
                                }

                        elif resource.resource_type == "worker":
                            worker = next((w for w in bop.workers if w.worker_id == resource.resource_id), None)
                            if worker:
                                resource_obj["name"] = worker.name
                                resource_obj["color"] = "#50c878"
                                resource_obj["size"] = {
                                    "width": 0.6,
                                    "height": 1.7,
                                    "depth": 0.6
                                }

                        elif resource.resource_type == "material":
                            material = next((m for m in bop.materials if m.material_id == resource.resource_id), None)
                            if material:
                                resource_obj["name"] = material.name
                                resource_obj["unit"] = material.unit
                                resource_obj["color"] = "#ffa500"
                                resource_obj["size"] = {
                                    "width": 0.5,
                                    "height": 0.3,
                                    "depth": 0.5
                                }

                        export_data["resources"].append(resource_obj)

        # JSON으로 변환
        json_str = json.dumps(export_data, indent=2, ensure_ascii=False)
        buffer = BytesIO(json_str.encode('utf-8'))
        buffer.seek(0)

        # 파일명 생성
        safe_filename = "".join(c for c in bop.project_title if c.isalnum() or c in (' ', '_', '-')).strip()
        if not safe_filename:
            safe_filename = "BOP"

        return StreamingResponse(
            buffer,
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={safe_filename}_3d.json"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"3D export 실패: {str(e)}")
